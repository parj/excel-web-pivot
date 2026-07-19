"""Generate sample_data/sales_sample.xlsx for testing the upload flow.

Sheet "Sales" has a two-row merged header (to exercise header flattening);
sheet "Employees" is a plain flat table. Run from the repo root:

    python backend/scripts/make_sample.py
"""
import datetime as dt
import pathlib
import random

from openpyxl import Workbook

out_dir = pathlib.Path(__file__).resolve().parents[2] / "sample_data"
out_dir.mkdir(exist_ok=True)

wb = Workbook()

# --- Sheet 1: Sales, with a merged two-row header ---
ws = wb.active
ws.title = "Sales"
ws.merge_cells("A1:A2"); ws["A1"] = "Date"
ws.merge_cells("B1:B2"); ws["B1"] = "Region"
ws.merge_cells("C1:C2"); ws["C1"] = "Product"
ws.merge_cells("D1:E1"); ws["D1"] = "Revenue"
ws["D2"] = "Units"; ws["E2"] = "Amount"
ws.merge_cells("F1:F2"); ws["F1"] = "Returned"

random.seed(42)
regions = ["East", "West", "North", "South"]
products = ["Widget", "Gadget", "Sprocket"]
for i in range(200):
    ws.append([
        dt.date(2025, 1, 1) + dt.timedelta(days=random.randint(0, 364)),
        random.choice(regions),
        random.choice(products),
        random.randint(1, 50),
        round(random.uniform(10, 2500), 2),
        random.random() < 0.08,
    ])

# --- Sheet 2: Employees, flat header ---
ws2 = wb.create_sheet("Employees")
ws2.append(["Name", "Department", "Hire Date", "Salary", "Active"])
depts = ["Sales", "Engineering", "Support", "Finance"]
names = ["Alex", "Sam", "Jordan", "Casey", "Riley", "Morgan", "Quinn", "Avery", "Drew", "Jamie"]
for i, name in enumerate(names):
    ws2.append([
        f"{name} {chr(65 + i)}.",
        random.choice(depts),
        dt.date(2018 + i % 7, (i % 12) + 1, (i % 27) + 1),
        random.randint(45000, 140000),
        i % 5 != 0,
    ])

path = out_dir / "sales_sample.xlsx"
wb.save(path)
print(f"Wrote {path}")
