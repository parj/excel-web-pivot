"""Excel parsing → schema inference → ClickHouse table creation + insert.

- .xlsx/.xlsm sheets are read via openpyxl so merged cells can be filled with
  their top-left value before header detection; .xls falls back to
  pandas/xlrd; .xlsb (the binary OOXML variant openpyxl can't open) is read
  via the xlsb_reader package.
- .xlsb has weaker fidelity than the other formats: xlsb_reader doesn't
  expose merged-cell ranges (so merged headers aren't filled) or pivot cache
  field names (so pivots are detected but not recreated as live configs —
  the underlying sheet is still imported as a table), and date cells come
  back as raw Excel serial numbers rather than datetimes.
- Multi-row headers are detected heuristically (leading rows that contain only
  text) and flattened into single names joined by " - ".
- Each sheet becomes a ReplacingMergeTree(_version) table keyed on _row_id;
  edits/deletes are versioned INSERTs, never mutations.
"""
from __future__ import annotations

import datetime as dt
import io
import json
import os
import re
import tempfile
import time
import uuid
import zipfile

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from xlsb_reader import XlsbWorkbook

from .config import settings
from .db import DB, ch

TYPE_TO_CH = {
    "number": "Nullable(Float64)",
    "date": "Nullable(DateTime64(3))",
    "bool": "Nullable(UInt8)",
    "string": "Nullable(String)",
}

BOOL_TOKENS = {"true": 1, "false": 0, "yes": 1, "no": 0}


class IngestError(Exception):
    pass


def check_file(filename: str, data: bytes):
    if len(data) > settings.max_upload_mb * 1024 * 1024:
        raise IngestError(f"File exceeds the {settings.max_upload_mb}MB limit")
    if not re.search(r"\.(xlsx|xlsm|xls|xlsb)$", filename, re.I):
        raise IngestError("Only .xlsx, .xlsm, .xls, and .xlsb files are supported")
    # Encrypted OOXML files (including .xlsb, which is a zip/OPC container
    # too) are wrapped in a CFB container (D0 CF 11 E0). A plain .xls is also
    # CFB, so only flag it for the zip-based extensions.
    if re.search(r"\.(xlsx|xlsm|xlsb)$", filename, re.I):
        if data[:4] == b"\xd0\xcf\x11\xe0":
            raise IngestError("This file appears to be password-protected; remove the password and re-upload")
        if not zipfile.is_zipfile(io.BytesIO(data)):
            raise IngestError("File is not a valid workbook (corrupt or wrong format)")


def _grids_from_wb(wb) -> dict[str, list[list]]:
    """Read every sheet as a raw 2D grid with merged cells filled."""
    grids = {}
    for ws in wb.worksheets:
        grid = [list(row) for row in ws.iter_rows(values_only=True)]
        for rng in ws.merged_cells.ranges:
            top_left = grid[rng.min_row - 1][rng.min_col - 1] if rng.min_row - 1 < len(grid) else None
            for r in range(rng.min_row - 1, min(rng.max_row, len(grid))):
                for c in range(rng.min_col - 1, min(rng.max_col, len(grid[r]))):
                    grid[r][c] = top_left
        grids[ws.title] = grid
    return grids


def _sheet_grid_xls(data: bytes) -> dict[str, list[list]]:
    frames = pd.read_excel(io.BytesIO(data), sheet_name=None, header=None, engine="xlrd")
    return {
        name: [[None if pd.isna(v) else v for v in row] for row in df.itertuples(index=False)]
        for name, df in frames.items()
    }


def _read_xlsb(data: bytes) -> tuple[dict[str, list[list]], list[dict]]:
    """Read a .xlsb workbook via xlsb_reader (openpyxl can't open the binary
    OOXML format). XlsbWorkbook only takes a path, so the upload is spooled
    to a temp file. Pivot tables are detected but not recreated as live
    configs — xlsb_reader doesn't expose pivot cache field names, so the
    underlying sheet is imported as a plain table instead (see module
    docstring for the other .xlsb fidelity gaps: no merged-cell fill, dates
    come back as raw Excel serial numbers)."""
    fd, path = tempfile.mkstemp(suffix=".xlsb")
    try:
        with os.fdopen(fd, "wb") as f:
            f.write(data)
        grids: dict[str, list[list]] = {}
        pivots: list[dict] = []
        with XlsbWorkbook(path) as wb:
            for sheet_name, values in wb.iter_values():
                if not values:
                    grids[sheet_name] = []
                    continue
                max_row = max(r for r, _ in values) + 1
                max_col = max(c for _, c in values) + 1
                grid = [[None] * max_col for _ in range(max_row)]
                for (r, c), v in values.items():
                    grid[r][c] = v
                grids[sheet_name] = grid
            for pt in wb.iter_pivot_tables():
                pivots.append({
                    "name": pt.get("name") or "PivotTable",
                    "sheet": pt.get("sheet"),
                    "error": "recreating pivot tables from .xlsb files isn't supported yet "
                             "— the underlying sheet was imported as a table instead",
                })
        return grids, pivots
    finally:
        os.unlink(path)


def _parse_source(filename: str, data: bytes) -> tuple[dict[str, list[list]], list[dict]]:
    """Dispatch to the right reader by extension and return (grids, pivots)."""
    name = filename.lower()
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        wb = load_workbook(io.BytesIO(data), data_only=True)
        return _grids_from_wb(wb), _extract_pivots(wb)
    if name.endswith(".xlsb"):
        return _read_xlsb(data)
    return _sheet_grid_xls(data), []


def _is_blank(v) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")


# ---------- pivot table detection (xlsx only) ----------

# Excel dataField subtotal -> our aggregation names.
PIVOT_AGG_MAP = {
    "sum": "sum",
    "count": "count",       # COUNTA
    "countNums": "count",
    "average": "avg",
    "max": "max",
    "min": "min",
}


def _resolve_pivot_source(wb, cache_source) -> str | None:
    """Source sheet name of a pivot cache: direct sheet ref or defined name."""
    if cache_source is None or cache_source.type != "worksheet":
        return None
    wss = cache_source.worksheetSource
    if wss is None:
        return None
    if wss.sheet:
        return wss.sheet
    if wss.name:  # named range, e.g. '=digit1'
        dn = wb.defined_names.get(wss.name.lstrip("="))
        if dn is not None:
            for sheet_title, _ref in dn.destinations:
                return sheet_title
    return None


def _extract_pivots(wb) -> list[dict]:
    """Pull every pivot table definition out of the workbook XML."""
    out = []
    for ws in wb.worksheets:
        for pt in getattr(ws, "_pivots", []):
            info: dict = {"name": getattr(pt, "name", None) or "PivotTable", "sheet": ws.title}
            try:
                info["ref"] = pt.location.ref
                info["source_sheet"] = _resolve_pivot_source(wb, pt.cache.cacheSource)
                fields = [f.name for f in pt.cache.cacheFields]
                info["rows"] = [fields[rf.x] for rf in pt.rowFields if 0 <= rf.x < len(fields)]
                info["columns"] = [fields[cf.x] for cf in pt.colFields if 0 <= cf.x < len(fields)]
                values, notes = [], []
                for df in pt.dataFields:
                    subtotal = df.subtotal or "sum"
                    agg = PIVOT_AGG_MAP.get(subtotal)
                    if agg is None:
                        notes.append(f"aggregation '{subtotal}' not supported, used sum")
                        agg = "sum"
                    values.append({"field_name": fields[df.fld], "agg": agg})
                info["values"] = values
                info["filters"] = [
                    fields[pf.fld] for pf in pt.pageFields if pf.fld is not None and 0 <= pf.fld < len(fields)
                ]
                info["notes"] = notes
            except Exception as e:
                info["error"] = str(e)
            out.append(info)
    return out


def _pivot_coverage(grid: list[list], refs: list[str]) -> float:
    """Fraction of a sheet's non-blank cells that fall inside pivot ranges."""
    try:
        bounds = [range_boundaries(r) for r in refs if r]
    except ValueError:
        return 0.0
    total = inside = 0
    for ri, row in enumerate(grid, start=1):
        for ci, v in enumerate(row, start=1):
            if _is_blank(v):
                continue
            total += 1
            if any(mc <= ci <= xc and mr <= ri <= xr for (mc, mr, xc, xr) in bounds):
                inside += 1
    return inside / total if total else 1.0


def _norm_header(s) -> str:
    return re.sub(r"\s+", " ", str(s)).strip().lower()


def _map_pivot_fields(pivot: dict, columns: list[dict]) -> dict | None:
    """Map cache-field header names to the ingested sheet's column names."""
    by_source = {_norm_header(c["source"]): c["name"] for c in columns}

    def m(header: str) -> str | None:
        return by_source.get(_norm_header(header))

    rows = [m(f) for f in pivot["rows"]]
    cols = [m(f) for f in pivot["columns"]]
    values = [{"field": m(v["field_name"]), "agg": v["agg"]} for v in pivot["values"]]
    filters = [{"field": m(f), "values": []} for f in pivot["filters"]]
    mapped = rows + cols + [v["field"] for v in values] + [f["field"] for f in filters]
    if any(x is None for x in mapped):
        return None
    return {"rows": rows, "columns": cols, "values": values, "filters": filters}


def _header_depth(grid: list[list]) -> int:
    """Consecutive leading rows containing only text → header rows (cap 3).
    If no typed (non-text) cell shows up in the first 10 rows, assume depth 1."""
    depth = 0
    for i, row in enumerate(grid[:10]):
        cells = [v for v in row if not _is_blank(v)]
        if cells and all(isinstance(v, str) for v in cells):
            depth += 1
            if depth == 3:
                break
        else:
            break
    return max(1, min(depth, 3)) if depth else 1


def _sanitize(name: str) -> str:
    s = re.sub(r"[^0-9a-zA-Z_]+", "_", name.strip()).strip("_")
    if not s:
        s = "col"
    if s[0].isdigit():
        s = "c_" + s
    return s[:80]


def _flatten_headers(grid: list[list], depth: int, width: int) -> list[dict]:
    cols = []
    seen: dict[str, int] = {}
    for c in range(width):
        parts = []
        for r in range(depth):
            v = grid[r][c] if r < len(grid) and c < len(grid[r]) else None
            if not _is_blank(v) and (not parts or str(v).strip() != parts[-1]):
                parts.append(str(v).strip())
        source = " - ".join(parts) if parts else f"Column {c + 1}"
        name = _sanitize(source)
        n = seen.get(name, 0)
        seen[name] = n + 1
        if n:
            name = f"{name}_{n + 1}"
        cols.append({"name": name, "source": source})
    return cols


def _infer_type(values: list) -> str:
    vals = [v for v in values if not _is_blank(v)]
    if not vals:
        return "string"
    if all(isinstance(v, bool) for v in vals):
        return "bool"
    if all(isinstance(v, str) and v.strip().lower() in BOOL_TOKENS for v in vals):
        return "bool"
    if all(isinstance(v, (dt.datetime, dt.date)) and not isinstance(v, bool) for v in vals):
        return "date"

    def numeric(v):
        if isinstance(v, bool):
            return False
        if isinstance(v, (int, float)):
            return True
        if isinstance(v, str):
            try:
                float(v.replace(",", ""))
                return True
            except ValueError:
                return False
        return False

    if sum(1 for v in vals if numeric(v)) >= 0.95 * len(vals):
        return "number"
    return "string"


def _coerce(v, typ: str):
    if _is_blank(v):
        return None
    try:
        if typ == "number":
            if isinstance(v, bool):
                return float(v)
            return float(str(v).replace(",", "")) if isinstance(v, str) else float(v)
        if typ == "bool":
            if isinstance(v, bool):
                return int(v)
            return BOOL_TOKENS.get(str(v).strip().lower())
        if typ == "date":
            if isinstance(v, dt.datetime):
                return v.replace(tzinfo=None)
            if isinstance(v, dt.date):
                return dt.datetime(v.year, v.month, v.day)
            return pd.to_datetime(v).to_pydatetime().replace(tzinfo=None)
    except (ValueError, TypeError):
        return None
    return str(v)


def coerce_value(v, typ: str):
    """Public coercion used by the row-edit endpoints."""
    return _coerce(v, typ)


def data_table_ddl(table_name: str, columns: list[dict]) -> str:
    col_defs = ",\n        ".join(f"`{c['name']}` {TYPE_TO_CH[c['type']]}" for c in columns)
    return f"""
    CREATE TABLE IF NOT EXISTS {DB}.`{table_name}` (
        _row_id String,
        _row_index UInt64,
        _version UInt64,
        _is_deleted UInt8 DEFAULT 0,
        {col_defs}
    ) ENGINE = ReplacingMergeTree(_version) ORDER BY _row_id
    """


def _prepare_sheet(sheet_name: str, grid: list[list], warnings: list[str]) -> tuple[list[dict], list[list]] | None:
    """Header-detect + type-infer a raw grid. Shared by fresh ingestion and
    by refresh (re-parsing the same sheet from a re-uploaded workbook)."""
    grid = [row for row in grid]
    # drop fully blank leading rows
    while grid and all(_is_blank(v) for v in grid[0]):
        grid.pop(0)
    if not grid or all(all(_is_blank(v) for v in r) for r in grid):
        warnings.append(f"Sheet '{sheet_name}' is empty — skipped")
        return None

    depth = _header_depth(grid)
    width = max(len(r) for r in grid)
    body = [r + [None] * (width - len(r)) for r in grid[depth:]]
    body = [r for r in body if not all(_is_blank(v) for v in r)]
    if len(body) > settings.max_rows_per_sheet:
        warnings.append(
            f"Sheet '{sheet_name}' has {len(body)} rows (limit {settings.max_rows_per_sheet}) — skipped"
        )
        return None

    columns = _flatten_headers(grid, depth, width)
    for ci, col in enumerate(columns):
        col["type"] = _infer_type([row[ci] for row in body[:2000]])
    return columns, body


def _ingest_sheet(client, workbook_id: str, sheet_name: str, grid: list[list], warnings: list[str]) -> dict | None:
    prepared = _prepare_sheet(sheet_name, grid, warnings)
    if prepared is None:
        return None
    columns, body = prepared
    width = len(columns)

    sheet_id = uuid.uuid4().hex
    table_name = f"data_{sheet_id}"
    client.command(data_table_ddl(table_name, columns))

    version = time.time_ns()
    rows = [
        [uuid.uuid4().hex, idx, version, 0] + [_coerce(row[ci], columns[ci]["type"]) for ci in range(width)]
        for idx, row in enumerate(body)
    ]
    col_names = ["_row_id", "_row_index", "_version", "_is_deleted"] + [c["name"] for c in columns]
    if rows:
        for start in range(0, len(rows), 50000):
            client.insert(f"{DB}.`{table_name}`", rows[start:start + 50000], column_names=col_names)

    client.insert(
        f"{DB}.sheets",
        [[sheet_id, workbook_id, sheet_name, table_name, json.dumps(columns), len(rows), dt.datetime.utcnow()]],
        column_names=["id", "workbook_id", "sheet_name", "table_name", "columns_json", "row_count", "created_at"],
    )
    return {"id": sheet_id, "sheet_name": sheet_name, "row_count": len(rows), "columns": columns}


def ingest_workbook(filename: str, data: bytes) -> dict:
    check_file(filename, data)
    try:
        grids, pivots = _parse_source(filename, data)
    except IngestError:
        raise
    except Exception as e:  # corrupt / unreadable
        raise IngestError(f"Could not parse workbook: {e}") from e

    client = ch()
    workbook_id = uuid.uuid4().hex
    sheets_meta, warnings = [], []

    # Sheets that are just rendered pivot output (>=80% of their non-blank
    # cells inside pivot ranges) are not ingested as tables — the pivots are
    # recreated live against their source sheets instead.
    refs_by_sheet: dict[str, list[str]] = {}
    for p in pivots:
        if "error" not in p:
            refs_by_sheet.setdefault(p["sheet"], []).append(p["ref"])
    pivot_only = {
        name for name, refs in refs_by_sheet.items() if _pivot_coverage(grids.get(name, []), refs) >= 0.8
    }

    for sheet_name, grid in grids.items():
        if sheet_name in pivot_only:
            continue
        meta = _ingest_sheet(client, workbook_id, sheet_name, grid, warnings)
        if meta:
            sheets_meta.append(meta)

    # Recreate detected pivots as live pivot configs on their source sheets.
    sheet_by_name = {m["sheet_name"]: m for m in sheets_meta}
    recreated: list[dict] = []
    failed_sheets: set[str] = set()
    for p in pivots:
        src = p.get("source_sheet")
        target = sheet_by_name.get(src) if src else None
        problem = p.get("error")
        cfg = None
        if problem is None and target is None:
            problem = f"source data (sheet '{src}') was not found in the workbook" if src else "source is external to the workbook"
        if problem is None:
            cfg = _map_pivot_fields(p, target["columns"])
            if cfg is None:
                problem = f"its fields did not match the columns of '{src}'"
        if problem is not None:
            warnings.append(f"Pivot '{p['name']}' (sheet '{p['sheet']}') was not recreated: {problem}")
            failed_sheets.add(p["sheet"])
            continue
        pivot_id = uuid.uuid4().hex
        now = dt.datetime.utcnow()
        client.insert(
            f"{DB}.pivot_configs",
            [[
                pivot_id, target["id"], p["name"],
                json.dumps(cfg["rows"]), json.dumps(cfg["columns"]),
                json.dumps(cfg["values"]), json.dumps(cfg["filters"]),
                0, now, now,
            ]],
            column_names=["id", "sheet_id", "name", "rows_json", "columns_json", "values_json",
                          "filters_json", "is_deleted", "created_at", "updated_at"],
        )
        for note in p.get("notes", []):
            warnings.append(f"Pivot '{p['name']}': {note}")
        recreated.append({"id": pivot_id, "name": p["name"], "sheet_id": target["id"], "source_sheet": src})

    # A pivot-only sheet whose pivots could not all be recreated still holds
    # data the user would otherwise lose — fall back to ingesting it as a table.
    for sheet_name in grids:
        if sheet_name in pivot_only and sheet_name in failed_sheets:
            meta = _ingest_sheet(client, workbook_id, sheet_name, grids[sheet_name], warnings)
            if meta:
                sheets_meta.append(meta)
        elif sheet_name in pivot_only:
            n = sum(1 for r in recreated if pivots and any(
                p["name"] == r["name"] and p["sheet"] == sheet_name for p in pivots))
            warnings.append(
                f"Sheet '{sheet_name}' contained only pivot output — recreated {n} live pivot(s) "
                f"from its source data instead of importing it as a table"
            )

    if not sheets_meta:
        raise IngestError("No usable sheets found in workbook. " + "; ".join(warnings))

    client.insert(
        f"{DB}.workbooks",
        [[workbook_id, filename, dt.datetime.utcnow()]],
        column_names=["id", "filename", "uploaded_at"],
    )
    return {
        "workbook_id": workbook_id,
        "filename": filename,
        "sheets": sheets_meta,
        "pivots": recreated,
        "warnings": warnings,
    }


# ---------- refresh (re-upload a source file to replace a workbook's data) ----------

# ALTER TABLE ... UPDATE runs as a background mutation by default; forcing it
# synchronous keeps the metadata read-your-writes consistent for the
# frontend, which reloads sheets right after the refresh job finishes.
_SYNC_MUTATION = {"mutations_sync": "1"}


def _tombstone_rows(client, table: str, columns: list[str]) -> None:
    live = client.query(f"SELECT _row_id, _row_index FROM {table} FINAL WHERE _is_deleted = 0").result_rows
    if not live:
        return
    version = time.time_ns()
    col_names = ["_row_id", "_row_index", "_version", "_is_deleted"] + columns
    tombstones = [[rid, ridx, version, 1] + [None] * len(columns) for rid, ridx in live]
    for start in range(0, len(tombstones), 50000):
        client.insert(table, tombstones[start:start + 50000], column_names=col_names)


def _clear_sheet(client, sheet: dict) -> None:
    """A previously-ingested sheet is missing from the refreshed file: wipe
    its rows but keep the sheet/table/pivots in place so nothing breaks."""
    table = f"{DB}.`{sheet['table_name']}`"
    _tombstone_rows(client, table, [c["name"] for c in sheet["columns"]])
    client.command(
        f"ALTER TABLE {DB}.sheets UPDATE row_count = 0 WHERE id = %(id)s",
        parameters={"id": sheet["id"]},
        settings=_SYNC_MUTATION,
    )


def _refresh_sheet(client, sheet: dict, grid: list[list], warnings: list[str]) -> dict | None:
    """Replace a sheet's data in place: existing rows are tombstoned and the
    freshly-parsed grid's rows are inserted new. The table/sheet id (and so
    any pivots built on top) is preserved. Columns are only ever widened —
    headers no longer present stay in the schema (as all-NULL going forward)
    rather than risk breaking a pivot config that still references them."""
    prepared = _prepare_sheet(sheet["sheet_name"], grid, warnings)
    if prepared is None:
        return None
    new_columns, body = prepared

    table = f"{DB}.`{sheet['table_name']}`"
    old_columns = sheet["columns"]
    old_names = {c["name"] for c in old_columns}
    added = [c for c in new_columns if c["name"] not in old_names]
    for c in added:
        client.command(f"ALTER TABLE {table} ADD COLUMN IF NOT EXISTS `{c['name']}` {TYPE_TO_CH[c['type']]}")
    merged_columns = old_columns + added
    merged_names = [c["name"] for c in merged_columns]
    # Existing columns keep coercing to their original (already-persisted)
    # type even if this refresh's data would infer differently — the
    # underlying ClickHouse column type never changes for them.
    type_by_name = {c["name"]: c["type"] for c in new_columns}
    type_by_name.update({c["name"]: c["type"] for c in old_columns})

    _tombstone_rows(client, table, merged_names)

    new_index_by_name = {c["name"]: i for i, c in enumerate(new_columns)}
    version = time.time_ns()
    rows = []
    for idx, row in enumerate(body):
        values = []
        for name in merged_names:
            ci = new_index_by_name.get(name)
            values.append(_coerce(row[ci], type_by_name[name]) if ci is not None else None)
        rows.append([uuid.uuid4().hex, idx, version, 0] + values)
    if rows:
        col_names = ["_row_id", "_row_index", "_version", "_is_deleted"] + merged_names
        for start in range(0, len(rows), 50000):
            client.insert(table, rows[start:start + 50000], column_names=col_names)

    client.command(
        f"ALTER TABLE {DB}.sheets UPDATE columns_json = %(cols)s, row_count = %(cnt)s WHERE id = %(id)s",
        parameters={"cols": json.dumps(merged_columns), "cnt": len(rows), "id": sheet["id"]},
        settings=_SYNC_MUTATION,
    )
    return {"id": sheet["id"], "sheet_name": sheet["sheet_name"], "row_count": len(rows), "columns": merged_columns}


def refresh_workbook(workbook_id: str, filename: str, data: bytes) -> dict:
    """Re-parse a re-uploaded source file and replace the data of a workbook
    that was already ingested, matching sheets by name. Table/sheet/pivot
    identities are preserved so existing pivots keep working against the
    refreshed data."""
    check_file(filename, data)
    client = ch()
    if not client.query(
        f"SELECT 1 FROM {DB}.workbooks WHERE id = %(id)s", parameters={"id": workbook_id}
    ).result_rows:
        raise IngestError("Workbook not found")

    try:
        grids, pivots = _parse_source(filename, data)
    except IngestError:
        raise
    except Exception as e:  # corrupt / unreadable
        raise IngestError(f"Could not parse workbook: {e}") from e

    sheets_res = client.query(
        f"SELECT id, sheet_name, table_name, columns_json FROM {DB}.sheets WHERE workbook_id = %(id)s",
        parameters={"id": workbook_id},
    )
    existing_by_name = {
        r[1]: {"id": r[0], "sheet_name": r[1], "table_name": r[2], "columns": json.loads(r[3])}
        for r in sheets_res.result_rows
    }

    # Same pivot-output detection as fresh ingestion, but only used to skip
    # newly-appeared pivot-output sheets — a sheet already tracked as a data
    # table keeps refreshing as one even if it now looks pivot-heavy, and we
    # never re-run the "recreate embedded pivots" step here (that would spawn
    # a duplicate live pivot config on every refresh).
    refs_by_sheet: dict[str, list[str]] = {}
    for p in pivots:
        if "error" not in p:
            refs_by_sheet.setdefault(p["sheet"], []).append(p["ref"])
    pivot_only = {
        name for name, refs in refs_by_sheet.items() if _pivot_coverage(grids.get(name, []), refs) >= 0.8
    }

    warnings: list[str] = []
    sheets_meta: list[dict] = []

    for sheet_name, grid in grids.items():
        existing_sheet = existing_by_name.pop(sheet_name, None)
        if existing_sheet is not None:
            meta_out = _refresh_sheet(client, existing_sheet, grid, warnings)
        elif sheet_name in pivot_only:
            continue
        else:
            meta_out = _ingest_sheet(client, workbook_id, sheet_name, grid, warnings)
            if meta_out:
                warnings.append(f"Sheet '{sheet_name}' is new — added to the workbook")
        if meta_out:
            sheets_meta.append(meta_out)

    # Sheets that existed before but are absent from the refreshed file keep
    # their table/pivots but lose their data.
    for sheet_name, existing_sheet in existing_by_name.items():
        _clear_sheet(client, existing_sheet)
        warnings.append(f"Sheet '{sheet_name}' was not found in the refreshed file — its data was cleared")

    client.command(
        f"ALTER TABLE {DB}.workbooks UPDATE refreshed_at = %(now)s WHERE id = %(id)s",
        parameters={"now": dt.datetime.utcnow(), "id": workbook_id},
        settings=_SYNC_MUTATION,
    )

    return {
        "workbook_id": workbook_id,
        "filename": filename,
        "sheets": sheets_meta,
        "warnings": warnings,
    }
